"""기능명세서(사용자용/개발자용)를 로컬 xlsx + 구글시트에 동시 반영.

사용법:
    .venv/bin/python scripts/push_feature_spec.py

- 데이터는 모두 이 파일의 DATA(사용자용), DEV_DATA(개발자용) 리스트에 있다.
  이 두 리스트만 수정하면 xlsx와 구글시트가 함께 갱신된다.
- 구글시트는 서비스 계정으로 인증 (KEY_FILE).
- 시트는 서비스 계정 이메일에 '편집자'로 공유돼 있어야 한다.
- 단방향(로컬 → 구글시트)이라 시트에서 직접 고친 내용은 다음 실행 때 덮어쓰여진다.
"""
from __future__ import annotations

import os

import gspread
from google.oauth2.service_account import Credentials
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ── 설정 ──────────────────────────────────────────────────────────────────────
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
KEY_FILE = os.path.join(BASE, "stock-screener-498805-c45880047de3.json")

# 사용자용·개발자용을 한 스프레드시트의 두 탭으로 관리
SHEET_ID = "1VhRygHUnN3VzTIrVG2B_WbKKpBxBQeXENs2FWp7t8Z4"

USER_XLSX_PATH = os.path.join(BASE, "기능명세서_사용자용.xlsx")
USER_WORKSHEET = "기능명세서(사용자용)"
USER_HEADERS = ["분류", "기능", "이런 기능이에요 (쉬운 설명)", "어디서 볼 수 있나요", "추가된 시점"]

# 새로 추가된 기능 — 사용자용 명세서에서 배경색으로 강조할 '기능' 라벨 집합.
# (기능 추가 시 여기에 라벨을 넣고, 다음 릴리스 때 비우는 식으로 관리)
NEW_USER_FEATURES = {"업종별 추천"}
NEW_FILL_HEX = "FFF3CD"  # 연한 노랑 (xlsx)
NEW_FILL_RGB = (1.0, 0.953, 0.804)  # 동일 색 (구글시트)

DEV_XLSX_PATH = os.path.join(BASE, "기능명세서.xlsx")
DEV_WORKSHEET = "기능명세서(개발자용)"
DEV_HEADERS = ["요구사항", "상세 기능", "데이터 정의", "엔터티 정의", "화면 정의", "패키지"]

# ── 사용자용 데이터 (분류, 기능, 쉬운 설명, 어디서 보나요, 추가된 시점) ──────────
DATA = [
    ("로그인", "로그인", "아이디와 비밀번호로 로그인하면 내 관심종목과 정보가 안전하게 보관됩니다.", "로그인 화면", "2026-05-26"),
    ("로그인", "로그아웃", "오른쪽 위 '로그아웃' 버튼으로 안전하게 빠져나옵니다.", "화면 오른쪽 위", "2026-06-02"),

    ("내 관심종목", "관심종목 담기", "관심 있는 회사를 별표(☆)를 눌러 목록에 담아둡니다.", "종목 화면의 별 버튼", "2026-05-26"),
    ("내 관심종목", "한눈에 보기", "담아둔 회사들의 지금 가격과 오르내림을 한 화면에 모아 보여줍니다.", "관심종목 화면", "2026-05-26"),
    ("내 관심종목", "실시간 가격", "장이 열려 있으면 가격이 자동으로 계속 바뀝니다. 새로고침할 필요가 없습니다.", "관심종목 화면", "2026-05-26"),
    ("내 관심종목", "정렬 바꾸기", "거래량이 많은 순서, 외국인이 많이 산 순서 등으로 줄을 바꿔 볼 수 있습니다.", "관심종목 위쪽 탭", "2026-05-26"),

    ("지금 시세 보기", "현재 가격", "지금 가격, 오늘 오른 정도, 오늘 최고가·최저가를 보여줍니다.", "종목 상세 화면", "2026-05-26"),
    ("지금 시세 보기", "1년 최고·최저", "지난 1년 동안 가장 비쌌던 가격과 가장 쌌던 가격을 알려줍니다.", "종목 상세 화면", "2026-05-26"),
    ("지금 시세 보기", "장 시간 외 거래", "아침 일찍이나 저녁 시간(8시~9시, 3시반~8시)에도 거래되는 가격을 자동으로 맞춰 보여줍니다.", "종목 상세 화면", "2026-06-02"),

    ("그래프 보기", "가격 그래프", "회사 가격이 그동안 어떻게 움직였는지 그래프로 보여줍니다. 기간(1주·1달·1년 등)을 골라 볼 수 있습니다.", "종목 상세 그래프", "2026-05-26"),
    ("그래프 보기", "평균선", "최근 가격의 평균을 선으로 그려, 흐름이 오르막인지 내리막인지 쉽게 보여줍니다.", "종목 상세 그래프", "2026-05-26"),
    ("그래프 보기", "살 만한 가격대 표시", "과거 거래가 많았던 가격대를 분석해, 사기 좋은 가격(지지선)과 팔기 좋은 가격(저항선)을 그래프에 선으로 그려 줍니다.", "종목 상세 그래프", "2026-06-02"),
    ("그래프 보기", "매수·매도 가격 안내", "지금 사면 좋은 가격 구간, 목표로 삼을 매도 가격, 손해를 끊을 가격을 숫자로 정리해 줍니다.", "매매 안내 카드", "2026-06-02"),
    ("그래프 보기", "유사 패턴 찾기", "지금 차트와 모양이 비슷했던 과거 사례를 다른 종목에서 찾아주고, 그때 이후에 얼마나 올랐는지/내렸는지를 알려줍니다. 비교 기준(DTW/피어슨/스피어만), 찾을 개수, 최소 유사도를 직접 정할 수 있습니다.", "유사 패턴 검색 카드", "2026-06-09"),
    ("그래프 보기", "유사 사례 차트 비교", "찾은 사례를 누르면 지금 종목과 그 사례의 실제 차트를 나란히 보여줍니다. 세로선(현재 시점) 왼쪽이 비슷했던 구간, 오른쪽이 그 뒤 실제로 흘러간 흐름입니다.", "유사 패턴 비교 차트", "2026-06-10"),

    ("AI 전망 분석", "종합 점수", "가격 흐름, 뉴스, 회사 재무를 모두 따져 이 회사가 좋은 신호인지 나쁜 신호인지 점수로 알려줍니다.", "전망 분석 화면", "2026-05-11"),
    ("AI 전망 분석", "뉴스·공시 해석", "AI가 최신 뉴스와 공시를 읽고, 호재인지 악재인지 쉬운 말로 요약해 줍니다.", "전망 분석 화면", "2026-05-11"),
    ("AI 전망 분석", "판단의 근거 보기", "점수가 왜 이렇게 나왔는지, 어떤 뉴스·공시를 보고 판단했는지 원문까지 확인할 수 있습니다.", "근거 목록 화면", "2026-05-11"),
    ("AI 전망 분석", "내 보유 종목 손익", "산 가격과 수량을 넣으면 지금 얼마 벌었는지/잃었는지, 본전이 되려면 얼마나 올라야 하는지 계산해 줍니다.", "보유 종목 카드", "2026-05-12"),

    ("종목 찾기", "업종별 추천", "반도체, 제약, 화학 등 29개 업종 중 하나를 누르면 그 업종에서 최근 가격과 거래량이 가장 좋은 종목 3개를 골라줍니다.", "업종 추천 화면", "2026-06-18"),
    ("종목 찾기", "조건으로 찾기", "'거래량이 갑자기 늘어난 회사', '외국인이 계속 사는 회사' 같은 조건을 골라, 그에 맞는 회사들을 찾아 줍니다.", "조건 검색 화면", "2026-05-27"),
    ("종목 찾기", "조건 세부 조정", "각 조건의 기준(며칠 연속인지, 몇 배인지 등)을 직접 바꿀 수 있습니다.", "조건 옆 설정 버튼", "2026-06-04"),
    ("종목 찾기", "조건 설명 보기", "어려운 조건 옆 ⓘ 버튼을 누르면 그 조건이 무슨 뜻인지 쉬운 말로 알려줍니다.", "조건 옆 ⓘ 버튼", "2026-06-09"),
    ("종목 찾기", "오늘의 인기 종목", "거래량·거래대금이 많은 회사, 많이 오른 회사, 외국인·기관이 많이 산 회사 순위를 보여줍니다.", "시장 현황 화면", "2026-05-26"),

    ("편하게 보기", "검색으로 찾기", "회사 이름이나 번호를 입력해 바로 찾아볼 수 있습니다.", "화면 위쪽 검색창", "2026-05-26"),
    ("편하게 보기", "살짝 미리보기", "목록에서 회사 위에 마우스를 올리면, 작은 그래프와 한 줄 요약을 옆에서 미리 보여줍니다.", "오른쪽 미리보기 칸", "2026-05-29"),
    ("편하게 보기", "뒤로가기", "회사를 보다가 뒤로가기를 누르면, 보던 목록의 그 자리로 그대로 돌아옵니다.", "전체 화면", "2026-06-02"),
    ("편하게 보기", "큰 글씨 모드", "버튼 하나로 글씨·그래프·여백까지 화면 전체를 1.3배 크게 키워, 작은 글씨가 불편한 분도 편하게 볼 수 있습니다.", "오른쪽 아래 큰 글씨 전환 버튼", "2026-06-01"),
]

# ── 개발자용 데이터 ───────────────────────────────────────────────────────────
# 대분류 행: ("1. 사용자 인증",) 처럼 1개짜리 튜플
# 기능 행:   (요구사항, 상세 기능, 데이터 정의, 엔터티 정의, 화면 정의, 패키지) 6개짜리 튜플
DEV_DATA = [
    ("1. 사용자 인증",),
    ("1-1. 관리자 로그인", "1-1-1-1. 아이디·비밀번호를 검증하고 세션 기반 로그인 토큰을 발급한다.", "로그인 토큰", "AuthToken", "로그인 화면", "web"),
    ("1-2. 세션 확인", "1-2-1-1. 진입 시 토큰 유효성을 확인하고 만료 시 로그인 화면으로 보낸다.", "로그인 토큰", "AuthToken", "", "web"),
    ("1-3. 로그아웃", "1-3-1-1. 토큰을 폐기하고 로그인 화면으로 이동한다.", "", "", "헤더 우측 로그아웃 버튼", "web"),

    ("2. 실시간 시세",),
    ("2-1. KIS WebSocket 단일 연결 유지", "2-1-1-1. KIS 실시간체결가 WebSocket에 단일 연결을 맺는다.\n2-1-1-2. 연결 끊김 시 자동 재연결한다.", "approval_key, 체결 틱", "KISConnectionManager", "", "services.realtime"),
    ("2-2. 종목별 구독 fan-out", "2-2-1-1. 브라우저 연결마다 asyncio.Queue를 생성한다.\n2-2-1-2. 수신한 틱을 해당 종목 구독 큐에 분배한다.\n2-2-1-3. 구독자가 0명인 종목은 KIS 구독을 해제한다.", "종목코드, 체결 틱", "TickQueue", "", "services.realtime"),
    ("2-3. KRX/NXT 시간대 자동 전환", "2-3-1-1. KST 기준 08:00~09:00, 15:30~20:00은 NXT, 그 외는 KRX 시장코드를 선택한다.", "시장 분류 코드", "MarketDivCode", "", "services.ranking"),
    ("2-4. 관심종목 다종목 시세 조회", "2-4-1-1. 최대 30종목의 현재가·등락률·거래량·거래대금을 일괄 조회한다.", "현재가, 등락률, 거래량, 거래대금", "WatchlistItem", "관심종목 목록 화면", "services.watchlist"),
    ("2-5. 단일 종목 현재가 시세", "2-5-1-1. 현재가, 시가/고가/저가, 거래량, 52주 최고·최저가를 조회한다.", "현재가, 시가, 고가, 저가, 거래량, 52주 최고/최저", "MarketQuote", "종목 상세 시세 영역", "services.position"),

    ("3. 관심종목 관리",),
    ("3-1. 관심종목 추가/해제", "3-1-1-1. 종목을 관심목록에 추가하거나 해제한다.", "종목코드 목록", "Watchlist", "관심종목 화면 · ☆ 버튼", "web"),
    ("3-2. 관심종목 정렬", "3-2-1-1. 거래량/거래대금/외국인/기관 기준으로 정렬한다.", "거래량, 거래대금, 순매수량", "WatchlistItem", "관심종목 정렬 탭", "services.watchlist"),
    ("3-3. 실시간 가격 반영", "3-3-1-1. WebSocket 틱으로 목록의 가격·등락률을 실시간 갱신한다.", "체결 틱", "WatchlistItem", "관심종목 목록 화면", "services.realtime"),

    ("4. 주식 전망 분석",),
    ("4-1. 정량(퀀트) 신호 산출", "4-1-1-1. 골든크로스, 이격도, 모멘텀, 거래량 급증, 외국인 수급 5종 신호를 계산한다.\n4-1-1-2. 각 신호를 방향·점수로 산출한다.", "이동평균, 이격도, 수익률, 거래량, 외국인 순매수", "QuantSignal", "전망 분석 신호 패널", "quant"),
    ("4-2. LLM 뉴스·공시 해석", "4-2-1-1. 수집된 뉴스·공시 텍스트를 LLM에 전달한다.\n4-2-1-2. 우선순위(실적·업황 > 수급 > 공시 이벤트)에 따라 방향(긍정/부정/중립)과 -8~+8 정수 점수, 요약, 신뢰도를 받는다.\n4-2-1-3. 동일 입력은 캐시로 중복 호출을 막는다.\n4-2-1-4. evidence에 없는 사실·추정은 사용하지 않도록 제약한다.", "뉴스, 공시 텍스트, 방향, 점수, 요약, 신뢰도", "AISignal", "전망 분석 LLM 요약 카드", "llm"),
    ("4-3. 재무 신호 산출", "4-3-1-1. ROE 등 재무 지표를 방향·점수로 평가한다.", "재무비율, 방향, 점수", "FinancialSignal", "전망 분석 신호 패널", "financial"),
    ("4-4. 종합 점수 합산", "4-4-1-1. 정량 합산, LLM 평균(±8), 재무 합산을 더해 종합 점수를 만든다.\n4-4-1-2. 점수 부호로 매수/매도/관망 방향을 결정한다.", "정량 점수, LLM 점수, 재무 점수, 종합 점수", "ScoreBreakdown", "최종 판단 카드", "analysis.scoring"),
    ("4-5. 근거(evidence) 추적", "4-5-1-1. 신호별로 참조한 뉴스·공시·재무 항목을 evidence_id로 연결한다.", "evidence_id, 출처, 제목, 원문 URL", "Evidence", "근거 목록 화면", "analysis"),
    ("4-6. 보유 종목 손익 분석", "4-6-1-1. 평단가·수량을 입력받아 평가손익, 손익분기율, 52주 고저 대비 거리를 계산한다.", "평단가, 수량, 평가손익, 손익분기율", "PositionContext", "보유 종목 컨텍스트 카드", "services.position"),
    ("4-7. ML 예측", "4-7-1-1. 학습된 모델로 상승/하락 확률과 기여 피처를 산출한다.", "예측 확률, 기여 피처", "MLPrediction", "ML 예측 카드", "ml"),

    ("5. 차트 분석",),
    ("5-1. 일봉 캔들 차트 표시", "5-1-1-1. 일봉 OHLCV와 이동평균(5/20/60/120)을 차트에 그린다.", "OHLCV, 이동평균", "OHLCVBar", "종목 상세 차트", "chart"),
    ("5-2. 오늘 캔들 실시간 갱신", "5-2-1-1. 실시간 틱으로 당일 미완성 캔들을 재생성 없이 갱신한다.", "체결 틱", "OHLCVBar", "종목 상세 차트", "chart"),
    ("5-3. 지지/저항 자동 탐지", "5-3-1-1. 일봉 거래량을 가격 구간에 분배해 Volume Profile을 만든다.\n5-3-1-2. POC와 HVN 존을 찾는다.\n5-3-1-3. 52주 고저·거래량 급증일 기준 Anchored VWAP을 계산한다.\n5-3-1-4. 지지선·저항선 가격대를 산출한다.", "거래량 프로파일, POC, HVN, VWAP", "SupportResistanceLevel", "차트 지지/저항선", "chart"),
    ("5-4. 매매 시그널 산출", "5-4-1-1. 지지/저항 기반 매수 구간·매도 목표·손절가·손익비를 계산한다.", "매수 구간, 매도 목표, 손절가, 손익비", "TradingSignal", "매매 시그널 카드", "chart"),
    ("5-5. 차트 패턴 유사 사례 검색", "5-5-1-1. 고른 구간 종가를 z-정규화한다.\n5-5-1-2. 전 종목(약 2,700) 슬라이딩 윈도우를 유클리드 거리로 1차 필터링한다.\n5-5-1-3. 선택 지표(DTW/피어슨/스피어만)로 정밀 점수·정렬한다.\n5-5-1-4. 결과 개수·최소 유사도(컷오프)를 적용한다.\n5-5-1-5. 각 사례의 이후 N일 수익률을 집계해 통계(평균·중앙값·상승비율)를 낸다.", "종가 시계열, 유사도 지표, 컷오프, 이후 수익률", "SimilarCase", "유사 패턴 검색 카드", "chart.pattern_match"),
    ("5-6. 유사 사례 실제 차트 비교", "5-6-1-1. 사례 선택 시 해당 종목 일봉을 조회해 유사 구간+이후 캔들을 그린다.\n5-6-1-2. 현재 종목 구간과 좌우로 나란히 비교 표시한다.\n5-6-1-3. 매칭 끝 시점에 '현재 시점' 세로선을 표시해 이후 흐름을 구분한다.", "일봉 OHLCV, 매칭 종료 시점", "OHLCVBar", "유사 패턴 비교 차트", "chart"),

    ("6. 업종별 추천",),
    ("6-1. 업종 목록 조회", "6-1-1-1. kospi_category.csv·kosdaq_category.csv에서 업종명을 stocks.sector 컬럼에 저장한다.\n6-1-1-2. daily_price 데이터가 있는 종목 기준으로 업종명·종목수를 집계해 반환한다.", "업종명, 종목수", "SectorInfo", "업종 추천 화면", "services.sector, services.screener_db"),
    ("6-2. 업종 내 top 3 추천", "6-2-1-1. 업종 내 종목을 모멘텀 스코어(20일 가격 모멘텀 × 0.6 + 거래량 서지 × 0.4)로 정렬한다.\n6-2-1-2. 상위 3종목의 종목코드·이름·현재가·등락률을 반환한다.", "종가 시계열, 거래량", "SectorPick", "업종 추천 결과 목록", "services.sector"),

    ("7. 다중 조건 검색식",),
    ("7-1. 조건 선택", "7-1-1-1. 28종 조건을 카테고리별 칩으로 다중 선택한다.", "조건 목록", "ScreenerCondition", "스크리너 조건 선택 화면", "web"),
    ("7-2. 조건 파라미터 설정", "7-2-1-1. 조건별 임계값(배수·일수·기간 등)을 칩 팝오버로 조정한다.", "임계값, 일수, 기간", "ScreenerParams", "조건 파라미터 팝오버", "web"),
    ("7-3. 조건 설명 제공", "7-3-1-1. 각 조건 칩의 ⓘ 버튼으로 쉬운 말 설명 팝오버를 표시한다.\n7-3-1-2. 설명 텍스트는 조건 정의에 정적으로 포함된다.", "조건 설명 텍스트", "ConditionItem", "조건 설명 팝오버", "web"),
    ("7-4. DB 기반 조건 평가", "7-4-1-1. 사전 수집된 일봉·투자자 데이터로 전체 종목을 검사한다.\n7-4-1-2. 거래량 급등·골든크로스·정배열·MACD·OBV 등 25종 조건을 계산한다.", "일봉 OHLCV, 투자자 데이터", "ScreenerResult", "스크리너 결과 목록", "services.screener_conditions"),
    ("7-5. 실시간 조건 평가", "7-5-1-1. 체결강도·신고가 근접·상한가 포착 3종은 KIS API로 실시간 조회한다.", "체결강도, 신고가, 상한가", "RankItem", "스크리너 결과 목록", "services.ranking"),
    ("7-6. 야간 데이터 수집 배치", "7-6-1-1. 전 종목 일봉 OHLCV와 투자자 데이터를 매일 수집해 SQLite에 저장한다.", "일봉 OHLCV, 투자자 데이터", "ScreenerDB", "", "services.screener_collector"),

    ("8. 시장현황 랭킹",),
    ("8-1. 거래량/거래대금 순위", "8-1-1-1. 거래량·거래대금 상위 종목을 조회한다.", "거래량, 거래대금 순위", "RankItem", "시장현황 랭킹 화면", "services.ranking"),
    ("8-2. 급등주 순위", "8-2-1-1. 등락률 상위 종목을 양수 등락률만 필터링해 조회한다.", "등락률 순위", "RankItem", "시장현황 랭킹 화면", "services.ranking"),
    ("8-3. 외국인/기관 순매수 순위", "8-3-1-1. 외국인·기관 순매수 상위 종목을 조회한다.", "순매수량 순위", "RankItem", "시장현황 랭킹 화면", "services.ranking"),
    ("8-4. 자동 갱신", "8-4-1-1. 실시간 탭은 15초마다 데이터를 갱신한다.", "순위 목록", "RankItem", "시장현황 랭킹 화면", "web"),

    ("9. 종목 탐색 · 화면 흐름",),
    ("9-1. 종목 검색", "9-1-1-1. 종목명·코드로 검색해 상세 페이지로 이동한다.", "종목명, 종목코드", "StockMaster", "헤더 검색창", "services.outlook"),
    ("9-2. 호버 미리보기", "9-2-1-1. 목록에서 종목 호버 시 미니 차트·한줄요약·시세를 우측 패널에 표시한다.", "미니 차트, 한줄요약, 시세", "MarketQuote", "우측 미리보기 패널", "web"),
    ("9-3. URL 라우팅 · 뒤로가기 복원", "9-3-1-1. 종목 진입 시 URL에 종목코드를 반영한다.\n9-3-1-2. 뒤로가기 시 직전 탭·스크롤 위치·스크리너 조건을 복원한다.", "종목코드, 화면 상태", "HistoryState", "", "web"),
    ("9-4. 종목 로고 표시", "9-4-1-1. 종목별 로고 이미지를 CDN에서 불러와 표시한다.", "로고 이미지", "StockLogo", "종목 목록·상세", "web"),

    ("10. 공통 · 환경",),
    ("10-1. 뉴스 수집", "10-1-1-1. 종목 관련 뉴스를 수집해 evidence로 만든다.", "뉴스 기사", "Evidence", "", "news"),
    ("10-2. 공시 수집", "10-2-1-1. DART 등에서 공시를 수집해 evidence로 만든다.", "공시 항목", "Evidence", "", "disclosure"),
    ("10-3. 큰 글씨 모드", "10-3-1-1. 토글로 전체 화면을 1.3배 확대한다.", "확대 배율", "ThemeState", "큰 글씨 토글", "web"),
    ("10-4. KIS 인증", "10-4-1-1. KIS Open API 접근 토큰과 WebSocket approval_key를 발급·관리한다.", "접근 토큰, approval_key", "KISAuth", "", "services"),
]

# ── 공통 스타일 상수 ──────────────────────────────────────────────────────────
_THIN = Side(style="thin", color="D0D0D0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(wrap_text=True, vertical="center")
_WRAP_LEFT = Alignment(wrap_text=True, vertical="top", horizontal="left")
_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")


def _auth_sheet():
    scopes = ["https://www.googleapis.com/auth/spreadsheets"]
    creds = Credentials.from_service_account_file(KEY_FILE, scopes=scopes)
    return gspread.authorize(creds).open_by_key(SHEET_ID)


def _push_table(sh, title: str, table: list[list[str]], header_rgb: tuple[float, float, float],
                col_widths: list[int] | None = None, highlight_rows: list[int] | None = None):
    ncols = len(table[0])
    try:
        ws = sh.worksheet(title)
        ws.clear()
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=title, rows=len(table) + 5, cols=ncols)
    ws.update(table, "A1")
    last_col = chr(64 + ncols)
    # 전체 셀 줄바꿈(wrap) 켜기 — 여러 줄 셀이 잘리지 않도록
    ws.format(f"A1:{last_col}{len(table)}", {"wrapStrategy": "WRAP", "verticalAlignment": "TOP"})
    ws.format(f"A1:{last_col}1", {
        "backgroundColor": {"red": header_rgb[0], "green": header_rgb[1], "blue": header_rgb[2]},
        "textFormat": {"bold": True, "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
        "horizontalAlignment": "CENTER",
        "verticalAlignment": "MIDDLE",
    })
    ws.freeze(rows=1)

    # 컬럼 폭(픽셀) 적용
    if col_widths:
        requests = [{
            "updateDimensionProperties": {
                "range": {"sheetId": ws.id, "dimension": "COLUMNS",
                          "startIndex": i, "endIndex": i + 1},
                "properties": {"pixelSize": w},
                "fields": "pixelSize",
            }
        } for i, w in enumerate(col_widths)]
        sh.batch_update({"requests": requests})

    # 신규 기능 행 배경색 강조 (1-based 시트 행번호 리스트)
    if highlight_rows:
        for row in highlight_rows:
            ws.format(f"A{row}:{last_col}{row}", {
                "backgroundColor": {"red": NEW_FILL_RGB[0], "green": NEW_FILL_RGB[1], "blue": NEW_FILL_RGB[2]},
            })


# ── 사용자용 ──────────────────────────────────────────────────────────────────
def build_user_xlsx() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = USER_WORKSHEET
    header_fill = PatternFill("solid", fgColor="548235")
    header_font = Font(bold=True, color="FFFFFF", size=13)
    cat_fill = PatternFill("solid", fgColor="E2EFDA")
    cat_font = Font(bold=True, size=12)
    body_font = Font(size=12)
    new_fill = PatternFill("solid", fgColor=NEW_FILL_HEX)

    ws.append(USER_HEADERS)
    for c in range(1, 6):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = _CENTER; cell.border = _BORDER
    ws.row_dimensions[1].height = 30

    r, prev_cat, cat_start = 2, None, 2
    for cat, func, desc, where, when in DATA:
        ws.append([cat, func, desc, where, when])
        is_new = func in NEW_USER_FEATURES
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.border = _BORDER; cell.font = body_font
            cell.alignment = _WRAP_LEFT if c == 3 else (_CENTER if c in (1, 4, 5) else _WRAP)
            if is_new and c >= 2:  # 분류(1열)는 카테고리 색 유지, 나머지에 신규 강조색
                cell.fill = new_fill
        if prev_cat is not None and cat != prev_cat:
            if r - 1 > cat_start:
                ws.merge_cells(start_row=cat_start, start_column=1, end_row=r - 1, end_column=1)
            cat_start = r
        prev_cat = cat
        r += 1
    if r - 1 > cat_start:
        ws.merge_cells(start_row=cat_start, start_column=1, end_row=r - 1, end_column=1)
    for row in range(2, r):
        cell = ws.cell(row=row, column=1)
        cell.fill = cat_fill; cell.font = cat_font; cell.alignment = _CENTER
        ws.row_dimensions[row].height = 42

    for col, w in zip("ABCDE", [16, 20, 66, 22, 16]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    wb.save(USER_XLSX_PATH)
    print(f"✓ 사용자용 로컬 저장: {USER_XLSX_PATH}")


# ── 개발자용 ──────────────────────────────────────────────────────────────────
def build_dev_xlsx() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = DEV_WORKSHEET
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    group_fill = PatternFill("solid", fgColor="D9E1F2")
    group_font = Font(bold=True, size=11)
    body_font = Font(size=11)

    ws.append(DEV_HEADERS)
    for c in range(1, 7):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = _CENTER; cell.border = _BORDER
    ws.row_dimensions[1].height = 24

    r = 2
    for entry in DEV_DATA:
        is_group = len(entry) == 1
        row = list(entry) + [""] * (6 - len(entry))
        ws.append(row)
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            cell.border = _BORDER
            cell.alignment = _WRAP_LEFT
            if is_group:
                cell.fill = group_fill; cell.font = group_font
            else:
                cell.font = body_font
        r += 1

    for col, w in zip("ABCDEF", [30, 52, 26, 20, 28, 22]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    wb.save(DEV_XLSX_PATH)
    print(f"✓ 개발자용 로컬 저장: {DEV_XLSX_PATH}")


def dev_table() -> list[list[str]]:
    table = [DEV_HEADERS]
    for entry in DEV_DATA:
        table.append([str(v) for v in (list(entry) + [""] * (6 - len(entry)))])
    return table


if __name__ == "__main__":
    build_user_xlsx()
    build_dev_xlsx()

    sh = _auth_sheet()
    user_table = [USER_HEADERS] + [list(r) for r in DATA]
    # 신규 기능 행 번호 (헤더가 1행이므로 DATA i번째 → 시트 i+2행)
    highlight = [i + 2 for i, r in enumerate(DATA) if r[1] in NEW_USER_FEATURES]
    _push_table(sh, USER_WORKSHEET, user_table, (0.33, 0.51, 0.21),
                col_widths=[110, 140, 460, 170, 110], highlight_rows=highlight)
    print(f"✓ 사용자용 탭 반영: {USER_WORKSHEET}")
    _push_table(sh, DEV_WORKSHEET, dev_table(), (0.18, 0.33, 0.59),
                col_widths=[210, 380, 190, 150, 200, 160])
    print(f"✓ 개발자용 탭 반영: {DEV_WORKSHEET}")
    print(f"→ https://docs.google.com/spreadsheets/d/{SHEET_ID}/edit")
    print("완료.")
